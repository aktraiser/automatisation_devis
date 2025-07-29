import streamlit as st
import pandas as pd
import pdfplumber
import google.generativeai as genai
import openpyxl
import json
import os
import tempfile
from io import BytesIO
from datetime import datetime
from dotenv import load_dotenv
import fitz  # PyMuPDF
import re
from PIL import Image
import logging
from typing import Dict, List, Optional, Any
import time
import hashlib
import io

# Imports pour OCR avec gestion d'erreur
try:
    import pytesseract
    TESSERACT_AVAILABLE = True
except ImportError:
    TESSERACT_AVAILABLE = False

# Configuration du logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Charger les variables d'environnement
load_dotenv()

# Configuration Streamlit
st.set_page_config(
    page_title="Extracteur Devis Simplifié", 
    layout="wide", 
    page_icon="🚗"
)

st.markdown("""
<style>
    .success-box { padding: 1rem; border-radius: 0.5rem; background-color: #d4edda; 
                   border: 1px solid #c3e6cb; color: #155724; margin: 1rem 0; }
    .warning-box { padding: 1rem; border-radius: 0.5rem; background-color: #fff3cd; 
                   border: 1px solid #ffeeba; color: #856404; margin: 1rem 0; }
    .stProgress > div > div > div > div { background-color: #4CAF50; }
</style>
""", unsafe_allow_html=True)

st.title("🚗 Extracteur de Devis Simplifié")

# Vérification de Tesseract
def verifier_tesseract():
    """Vérifie si Tesseract est disponible"""
    if not TESSERACT_AVAILABLE:
        return False, "pytesseract non installé"
    
    try:
        version = pytesseract.get_tesseract_version()
        return True, str(version)
    except Exception as e:
        return False, str(e)

class DevisExtractorSimplifie:
    """Extracteur de devis simplifié avec OCR pour PDFs image/texte"""
    
    def __init__(self, api_key: str):
        self.api_key = api_key
        genai.configure(api_key=api_key)
        self.model = genai.GenerativeModel('gemini-2.0-flash')
        self.ocr_disponible, self.tesseract_version = verifier_tesseract()
        
    def extraire_texte_avec_ocr(self, fichier_pdf_bytes: bytes, filename: str) -> dict:
        """Extraction de texte avec OCR pour les PDFs image/texte"""
        st.info(f"🔄 Extraction pour {filename}")
        
        try:
            if not fichier_pdf_bytes or len(fichier_pdf_bytes) == 0:
                return {"succes": False, "erreur": "Fichier PDF vide"}
            
            doc = fitz.open(stream=fichier_pdf_bytes, filetype="pdf")
            
            pages_info = []
            texte_complet = ""
            nombre_pages = len(doc)
            
            if nombre_pages > 1:
                progress = st.progress(0)
                status = st.empty()
            
            for num_page in range(nombre_pages):
                if nombre_pages > 1:
                    status.text(f"Analyse page {num_page + 1}/{nombre_pages}")
                    
                page = doc.load_page(num_page)
                
                # D'abord essayer l'extraction directe
                texte_page = page.get_text()
                methode_extraction = "direct"
                
                # Si pas de texte, essayer l'OCR sur les images
                if not texte_page.strip():
                    texte_page = ""
                    images = page.get_images()
                    
                    if images and self.ocr_disponible and TESSERACT_AVAILABLE:
                        methode_extraction = "OCR"
                        st.info(f"🔍 OCR nécessaire page {num_page + 1} - {len(images)} image(s)")
                        
                        for img_index, img in enumerate(images):
                            try:
                                xref = img[0]
                                base_image = doc.extract_image(xref)
                                image_bytes = base_image["image"]
                                image = Image.open(io.BytesIO(image_bytes))
                                
                                texte_img = pytesseract.image_to_string(image, lang='fra')
                                
                                if texte_img.strip():
                                    texte_page += f"\n--- Image {img_index + 1} ---\n{texte_img}\n"
                                    
                            except Exception as e:
                                st.warning(f"⚠️ Erreur OCR image {img_index + 1}, page {num_page + 1}: {e}")
                                continue
                        
                        if texte_page.strip():
                            st.success(f"✅ OCR réussi page {num_page + 1}: {len(texte_page)} caractères")
                    
                    elif not self.ocr_disponible and images:
                        st.warning(f"⚠️ Images détectées page {num_page + 1} mais OCR non disponible")
                
                info_page = {
                    "numero_page": num_page + 1,
                    "nombre_caracteres": len(texte_page),
                    "methode": methode_extraction,
                    "images_detectees": len(page.get_images())
                }
                
                pages_info.append(info_page)
                texte_complet += f"\n--- Page {num_page + 1} ---\n{texte_page}\n"
                
                if nombre_pages > 1:
                    progress.progress((num_page + 1) / nombre_pages)
            
            if nombre_pages > 1:
                progress.empty()
                status.empty()
            
            doc.close()
            
            return {
                "succes": True,
                "nombre_pages": nombre_pages,
                "pages_info": pages_info,
                "texte_complet": self._normaliser_texte(texte_complet),
                "taille_totale": len(texte_complet),
                "ocr_utilise": any(p["methode"] == "OCR" for p in pages_info)
            }
            
        except Exception as e:
            logger.error(f"Erreur extraction OCR {filename}: {str(e)}")
            return {"succes": False, "erreur": str(e)}
    
    def _normaliser_texte(self, texte: str) -> str:
        """Normalise le texte extrait"""
        if not texte:
            return texte
        
        texte = re.sub(r'\s+', ' ', texte)
        texte = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F-\x9F]', ' ', texte)
        
        # Correction des ligatures communes des PDFs
        ligatures = {
            "ﬀ": "ff", "ﬁ": "fi", "ﬂ": "fl", "ﬃ": "ffi", "ﬄ": "ffl",
            "ﬅ": "ft", "ﬆ": "st", "Æ": "AE", "æ": "ae", "œ": "oe", "Œ": "OE"
        }
        for search, replace in ligatures.items():
            texte = texte.replace(search, replace)
        
        return texte.strip()
    
    def get_expert_prompt(self) -> str:
        """Prompt d'expert pour extraction de devis automobile"""
        
        return """Tu es un expert en devis automobiles avec 20 ans d'expérience. Tu analyses TOUS types de documents automobiles : devis, simulations, offres commerciales.

MISSION : Extraire TOUS les devis/simulations présents dans le document.

INFORMATIONS À CHERCHER :

🚗 **VÉHICULE** : Marque + Modèle complet (BMW iX1, Mercedes GLC, etc.)
🏢 **LOUEUR** : BMW Finance, Arval, Alphabet, Ayvens, Athlon, concessionnaires
🔖 **RÉFÉRENCE LEASEUR** : Numéro de référence ou code du devis/contrat
💰 **LOYERS V1** : Le montant mensuel TOTAL que paie le client
💳 **PRIX CATALOGUE** : Prix du véhicule neuf avec options
💎 **PRIX OPTIONS** : Prix des options/accessoires seulement
🔋 **PRIX BATTERIE** : Prix de la batterie électrique (si applicable)
💰 **1ER LOYER** : Montant du premier loyer (si différent des autres)
⚡ **PUISSANCE** : Puissance du véhicule (ex: "286CH-6cv")
📋 **LOI DE ROULAGE** : Durée en mois + kilométrage (format "39/70")
📅 **DATE OFFRE** : Date de l'offre (JJ/MM/AAAA)
🛡️ **PRESTATIONS** : Services inclus (assurance, entretien, assistance)
📊 **AUTRES** : TVS, CO2, consommation mixte (L/100km, kWh/100km ou Wh/km)
💼 **AND** : Amortissements non déductibles / mois (en euros)

RÈGLES :
1. Chaque véhicule différent = 1 devis séparé
2. Si plusieurs montants mensuels : prends le PLUS ÉLEVÉ pour loyers_v1
3. Adapte-toi à TOUS les formats de loueurs
4. Sois tolérant aux erreurs OCR

RETOURNE EXCLUSIVEMENT UN JSON :
{
  "devis_trouves": [
    {
      "vehicule": "VOLKSWAGEN ID 4 LIFE MAX ELECTRIQUE 286 CH",
      "loueur": "VOLKSWAGEN BANK",
      "reference_leaseur": "VW-2025-001234",
      "date_offre": "02/05/2025",
      "prix_catalogue": 48063.76,
      "prix_options": 2850.00,
      "prix_batterie": 8500.00,
      "premier_loyer": 750.00,
      "puissance": "286CH-6cv",
      "co2": 0,
      "consommation": "152.80 Wh/km",
      "loi_de_roulage": "37/80",
      "presta": "Maintenance, Assistance, Perte financière",
      "loyers_v1": 667.83,
      "tvs": 720.00,
      "and": 666.00,
      "numero_devis_dans_pdf": 1
    }
  ],
  "nombre_devis_detectes": 1
}"""

    def extract_with_llm(self, texte_complet: str, filename: str, info_extraction: dict) -> List[dict]:
        """Extraction avec LLM - retourne TOUS les devis trouvés"""
        
        try:
            contexte_ocr = ""
            if info_extraction.get("ocr_utilise"):
                contexte_ocr = f"CONTEXTE OCR : Le texte peut contenir des erreurs de reconnaissance OCR"
            
            prompt = f"""{self.get_expert_prompt()}

Analyse ce document de devis automobile et extrais TOUS les devis présents.

FICHIER: {filename}
{contexte_ocr}

TEXTE DU DOCUMENT:
'''{texte_complet[:25000]}'''
"""
            
            # Appel Gemini avec retry
            max_retries = 3
            for attempt in range(max_retries):
                try:
                    response = self.model.generate_content(
                        prompt,
                        generation_config=genai.types.GenerationConfig(
                            temperature=0.1,
                            max_output_tokens=100000,
                            response_mime_type="application/json"
                        ),
                        request_options={"timeout": 120}
                    )
                    break
                except Exception as e:
                    if attempt < max_retries - 1:
                        wait_time = (2 ** attempt) * 1
                        logger.warning(f"Tentative {attempt + 1} échouée, retry dans {wait_time}s: {str(e)}")
                        time.sleep(wait_time)
                    else:
                        raise
            
            if not response.text:
                st.warning(f"⚠️ Aucune extraction pour {filename}")
                return [{"vehicule": os.path.splitext(filename)[0], "loueur": "EXTRACTION_ECHOUEE"}]
            
            try:
                result = json.loads(response.text)
                
                if isinstance(result, list):
                    devis_list = result
                elif isinstance(result, dict):
                    devis_list = result.get("devis_trouves", [])
                else:
                    st.warning(f"⚠️ Format JSON inattendu")
                    devis_list = []
                
                st.success(f"✅ {len(devis_list)} devis extraits pour {filename}")
                
            except json.JSONDecodeError as json_err:
                logger.error(f"Erreur JSON pour {filename}: {json_err}")
                st.error(f"⚠️ Erreur JSON pour {filename}")
                return [{"vehicule": os.path.splitext(filename)[0], "loueur": "ERREUR_JSON"}]
            
            # Post-traitement
            processed_devis = []
            for i, devis in enumerate(devis_list):
                processed_devis.append(self._post_process_extraction(devis, filename, i+1))
            
            return processed_devis if processed_devis else [{"vehicule": os.path.splitext(filename)[0], "loueur": "AUCUN_DEVIS"}]
            
        except Exception as e:
            logger.error(f"Erreur extraction LLM pour {filename}: {str(e)}")
            st.error(f"⚠️ Erreur LLM pour {filename}: {str(e)}")
            return [{"vehicule": os.path.splitext(filename)[0], "loueur": "ERREUR_API"}]
    
    def _post_process_extraction(self, result: dict, filename: str, numero_devis: int = 1) -> dict:
        """Post-traitement des données extraites"""
        
        result.setdefault("vehicule", f"{os.path.splitext(filename)[0]} - Devis {numero_devis}")
        result.setdefault("loueur", "NON_IDENTIFIE")
        result["fichier_source"] = filename
        
        # Conversion AND annuel → mensuel si nécessaire
        if result.get("and") and result["and"] > 1000:
            result["and"] = round(result["and"] / 12, 2)
        
        # Nettoyage des valeurs numériques
        numeric_fields = ["prix_catalogue", "prix_options", "prix_batterie", "premier_loyer", "co2", "tvs", "and", "loyers_v1"]
        for field in numeric_fields:
            if field in result and result[field] is not None:
                try:
                    result[field] = float(result[field])
                except (ValueError, TypeError):
                    result[field] = None
        
        return result

def process_pdfs_simplifie(extractor: DevisExtractorSimplifie, pdfs: List[Any]) -> List[dict]:
    """Traite les PDFs avec l'extracteur simplifié"""
    
    all_records = []
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    for i, pdf in enumerate(pdfs):
        status_text.text(f"Traitement {i+1}/{len(pdfs)}: {pdf.name}")
        
        try:
            raw_bytes = pdf.read()
            
            if len(raw_bytes) == 0:
                st.error(f"❌ Fichier vide : {pdf.name}")
                all_records.append({
                    "vehicule": os.path.splitext(pdf.name)[0],
                    "loueur": "FICHIER_VIDE",
                    "fichier_source": pdf.name
                })
                continue
            
            # Extraction du texte
            info_extraction = extractor.extraire_texte_avec_ocr(raw_bytes, pdf.name)
            
            if not info_extraction["succes"]:
                st.error(f"❌ Extraction échouée pour {pdf.name}")
                all_records.append({
                    "vehicule": os.path.splitext(pdf.name)[0],
                    "loueur": "ERREUR_EXTRACTION",
                    "fichier_source": pdf.name
                })
                continue
            
            st.success(f"✅ Extraction réussie pour {pdf.name}")
            
            # Extraction avec LLM
            devis_list = extractor.extract_with_llm(
                info_extraction["texte_complet"], 
                pdf.name, 
                info_extraction
            )
            
            all_records.extend(devis_list)
            
        except Exception as e:
            logger.error(f"Erreur traitement {pdf.name}: {str(e)}")
            st.error(f"❌ Erreur {pdf.name}: {str(e)}")
            all_records.append({
                "vehicule": os.path.splitext(pdf.name)[0],
                "loueur": "ERREUR_TRAITEMENT",
                "fichier_source": pdf.name
            })
        
        progress_bar.progress((i + 1) / len(pdfs))
    
    status_text.empty()
    progress_bar.empty()
    
    return all_records

def export_excel_suivi(df: pd.DataFrame) -> BytesIO:
    """Crée un Excel de suivi classique (format tableau)"""
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Suivi Devis"
    
    # Headers
    headers = [
        "Véhicule", "Loueur", "Prix Catalogue", "Prix Options", "Prix Batterie", "Loyer Mensuel", 
        "Loi de Roulage", "Date Offre", "CO2", "Consommation", "TVS", "AND", 
        "Prestations", "Fichier Source", "Remarques"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
    
    # Données
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        ws.cell(row=row_idx, column=1, value=row.get('vehicule', ''))
        ws.cell(row=row_idx, column=2, value=row.get('loueur', ''))
        ws.cell(row=row_idx, column=3, value=row.get('prix_catalogue', ''))
        ws.cell(row=row_idx, column=4, value=row.get('prix_options', ''))
        ws.cell(row=row_idx, column=5, value=row.get('prix_batterie', ''))
        ws.cell(row=row_idx, column=6, value=row.get('loyers_v1', ''))
        ws.cell(row=row_idx, column=7, value=row.get('loi_de_roulage', ''))
        ws.cell(row=row_idx, column=8, value=row.get('date_offre', ''))
        ws.cell(row=row_idx, column=9, value=row.get('co2', ''))
        ws.cell(row=row_idx, column=10, value=row.get('consommation', ''))
        ws.cell(row=row_idx, column=11, value=row.get('tvs', ''))
        ws.cell(row=row_idx, column=12, value=row.get('and', ''))
        ws.cell(row=row_idx, column=13, value=row.get('presta', ''))
        ws.cell(row=row_idx, column=14, value=row.get('fichier_source', ''))
        ws.cell(row=row_idx, column=15, value=row.get('remarques', ''))
    
    # Ajuster les largeurs
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def inject_in_excel_suivi(df: pd.DataFrame, template_wb: openpyxl.Workbook) -> BytesIO:
    """Injection dans un Excel de suivi existant"""
    
    ws = template_wb.active
    headers = [c.value for c in ws[1]]
    
    col_idx = {h: i + 1 for i, h in enumerate(headers)}
    
    # Mapping simple pour injection dans Excel de suivi
    mapping = {
        "vehicule": "VEHICULE",
        "prix_catalogue": "PRIX CATALOGUE\nVéhicule + option",
        "prix_options": "Prix Options",
        "prix_batterie": "Prix Batterie",
        "premier_loyer": "1er Loyer",
        "co2": "Rejet de CO2",
        "consommation": "conso",
        "tvs": "TVS",
        "and": "AND",
        "loi_de_roulage": "Loi de roulage",
        "presta": "Presta",
        "loueur": "LOUEUR",
        "date_offre": "DATE dernière OFFRE",
        "loyers_v1": "LOYERS\nV1",
        "remarques": "REMARQUE"
    }
    
    # Mapping par position de colonne si les headers ne correspondent pas exactement
    position_mapping = {
        "consommation": 3,  # Colonne C
        "co2": 4,          # Colonne D  
        "prix_batterie": 5, # Colonne E
        "tvs": 6,          # Colonne F
        "and": 7           # Colonne G
    }
    
    inserted_count = 0
    for row_idx, row in df.iterrows():
        next_row = ws.max_row + 1
        
        for key, val in row.items():
            if val is not None and str(val).strip():
                col_num = None
                
                # Essayer d'abord le mapping par position de colonne
                if key in position_mapping:
                    col_num = position_mapping[key]
                
                # Sinon, utiliser le mapping par nom de header
                elif key in mapping:
                    header = mapping[key]
                    if header in col_idx:
                        col_num = col_idx[header]
                
                if col_num:
                    # Traitement spécial pour la consommation
                    if key == "consommation":
                        val_str = str(val).strip()
                        # Si c'est juste un nombre, essayer d'ajouter l'unité appropriée
                        if val_str.replace('.', '').replace(',', '').isdigit():
                            val_num = float(val_str.replace(',', '.'))
                            # Détection automatique d'unité selon la valeur
                            if val_num > 50:  # Probablement Wh/km (ex: 152.80)
                                val = f"{val_num} Wh/km"
                            elif val_num > 10:  # Probablement kWh/100km (ex: 16.8)
                                val = f"{val_num} kWh/100km"
                            else:  # Probablement L/100km (ex: 6.5)
                                val = f"{val_num} L/100km"
                        # Si déjà formaté, garder tel quel
                        ws.cell(row=next_row, column=col_num).value = val
                    else:
                        if isinstance(val, (int, float)):
                            val = round(val, 2)
                        
                        ws.cell(row=next_row, column=col_num).value = val
        
        inserted_count += 1
    
    st.success(f"✅ {inserted_count} devis injectés dans le template Excel")
    
    output = BytesIO()
    template_wb.save(output)
    output.seek(0)
    return output

def inject_in_matrice_tco(df: pd.DataFrame, template_wb: openpyxl.Workbook) -> BytesIO:
    """Injection dans une matrice TCO existante (format transposé)"""
    
    ws = template_wb.active
    
    # Lire les champs de la matrice (colonne A ou B selon le format)
    champs_excel = {}
    
    for row in range(1, min(ws.max_row + 1, 50)):  # Limiter la recherche
        for col in [1, 2]:  # Chercher en colonne A et B
            cell_value = ws.cell(row=row, column=col).value
            if cell_value and str(cell_value).strip():
                champ_name = str(cell_value).strip()
                if any(keyword in champ_name.lower() for keyword in ['véhicule', 'prix', 'loyer', 'co2', 'loueur']):
                    champs_excel[champ_name] = row
    
    # Trouver les colonnes pour injection (D, F, H, J, L, N, P... - éviter C, E, G, I, K, M, O)
    start_col = 4  # Colonne D = 4
    max_vehicules = min(len(df), 10)
    
    # Mapping pour matrice TCO
    mapping_matrice = {
        "marque": ["Marque", "MARQUE", "marque"],
        "modele": ["Modèle", "MODELE", "modele", "Véhicule"],
        "loueur": ["Loueur", "LOUEUR", "loueur"],
        "prix_catalogue": ["Prix catalogue", "PRIX CATALOGUE", "Prix", "Véhicule + option"],
        "prix_options": ["Prix options", "options", "accessoires", "Options/Accessoires"],
        "prix_batterie": ["Prix batterie", "batterie", "Batterie électrique"],
        "premier_loyer": ["1er loyer", "Premier loyer", "First payment"],
        "loyers_v1": ["Loyer", "LOYER", "Loyer mensuel", "LOYERS", "V1"],
        "co2": ["CO2", "co2", "Rejet CO2"],
        "consommation": ["Consommation", "consommation", "Consommation mixte", "conso"],
        "tvs": ["TVS", "tvs"],
        "and": ["AND", "and", "Amortissement"],
        "date_offre": ["Date", "DATE", "Date offre"]
    }
    
    # Injection véhicule par véhicule
    for idx, (_, vehicule_data) in enumerate(df.iterrows()):
        if idx >= max_vehicules:
            break
            
        # Colonnes D(4), F(6), H(8), J(10), L(12), N(14), P(16)...
        target_col = start_col + (idx * 2)
        col_letter = openpyxl.utils.get_column_letter(target_col)
        # Traitement spécial pour le véhicule → séparer en marque, modèle et finition
        vehicule_complet = vehicule_data.get('vehicule', '')
        if vehicule_complet and str(vehicule_complet).strip():
            # Séparer marque, modèle et finition
            parts = str(vehicule_complet).strip().split(' ')
            
            # Marque = Premier mot
            marque = parts[0] if parts else ''
            
            # Modèle = Deuxième mot (si disponible)
            modele = parts[1] if len(parts) > 1 else ''
            
            # Finition = Tout le reste (à partir du 3ème mot)
            finition = ' '.join(parts[2:]) if len(parts) > 2 else ''
            
            # Cas spécial : si pas de 3ème mot, utiliser l'ancienne logique avec finitions connues
            if not finition and len(parts) > 2:
                finitions_connues = ['Clever', 'Style', 'Design', 'Life', 'Pro', 'Sport', 'Comfort', 'Edition', 'Plus', 
                                   'Max', 'Executive', 'Business', 'Premium', 'Lounge', 'Selection', 'Advance', 
                                   'Active', 'Dynamic', 'Elegance', 'Excellence', 'S-Line', 'R-Line', 'E-Tech', 'Techno']
                
                # Vérifier si le dernier mot est une finition connue
                if parts[-1] in finitions_connues:
                    finition = parts[-1]
                    # Le modèle devient tout sauf la marque et la finition
                    modele = ' '.join(parts[1:-1]) if len(parts) > 2 else parts[1]
            
            # Injecter marque (ligne 6)
            if marque:
                try:
                    ws.cell(row=5, column=target_col).value = marque
                except Exception as e:
                    st.warning(f"      ⚠️ Erreur marque: {e}")
            
            # Injecter modèle (ligne 6)  
            if modele:
                try:
                    ws.cell(row=6, column=target_col).value = modele
                except Exception as e:
                    st.warning(f"      ⚠️ Erreur modèle: {e}")
            
            # Injecter finition (ligne 7)
            if finition:
                try:
                    ws.cell(row=7, column=target_col).value = finition
                except Exception as e:
                    st.warning(f"      ⚠️ Erreur finition: {e}")
        
        # Injection référence leaseur (ligne 9)
        reference_leaseur = vehicule_data.get('reference_leaseur', '') or vehicule_data.get('ref_leaseur', '') or vehicule_data.get('reference', '')
        if reference_leaseur and str(reference_leaseur).strip() not in ['', 'nan', 'None']:
            try:
                ws.cell(row=9, column=target_col).value = str(reference_leaseur).strip()
            except Exception as e:
                st.warning(f"      ⚠️ Erreur référence leaseur L9: {e}")
        
        
        # Injection prix batterie (ligne 21)
        prix_batterie = vehicule_data.get('prix_batterie', '') or vehicule_data.get('batterie', '') or vehicule_data.get('prix_batterie_electrique', '')
        if prix_batterie and str(prix_batterie).strip() not in ['', 'nan', 'None']:
            try:
                prix_batterie_value = float(prix_batterie)
                ws.cell(row=21, column=target_col).value = prix_batterie_value
            except (ValueError, TypeError):
                # Si ce n'est pas un nombre, injecter tel quel
                ws.cell(row=21, column=target_col).value = str(prix_batterie)
        
        # Injection loyer mensuel (ligne 38)
        loyer_mensuel = vehicule_data.get('loyers_v1', '') or vehicule_data.get('loyer', '') or vehicule_data.get('loyer_mensuel', '')
        if loyer_mensuel and str(loyer_mensuel).strip() not in ['', 'nan', 'None']:
            try:
                loyer_value = float(loyer_mensuel)
                ws.cell(row=38, column=target_col).value = loyer_value
            except (ValueError, TypeError):
                # Si ce n'est pas un nombre, injecter tel quel
                ws.cell(row=38, column=target_col).value = str(loyer_mensuel)
        
        # Traitement spécial pour durée, kilométrage et consommation
        # Loi de roulage (format "39/70" = 39 mois / 70 000 km)
        loi_roulage = str(vehicule_data.get('loi_de_roulage', '')).strip()
        if loi_roulage and '/' in loi_roulage:
            parts = loi_roulage.split('/', 1)
            duree = parts[0].strip() if parts else ''
            km = parts[1].strip() if len(parts) > 1 else ''
            
            # Injection durée (ligne 25)
            if duree:
                try:
                    # Injecter seulement le nombre, sans "mois"
                    duree_value = int(duree) if duree.isdigit() else duree
                    ws.cell(row=25, column=target_col).value = duree_value
                except Exception as e:
                    st.warning(f"      ⚠️ Erreur durée L25: {e}")
            
            # Injection kilométrage (ligne 26) 
            if km:
                try:
                    # Injecter seulement le nombre en milliers, sans "km"
                    if km.isdigit():
                        km_value = int(km) * 1000  # Convertir en valeur complète
                    else:
                        # Essayer d'extraire juste le nombre si déjà formaté
                        km_value = km.replace('km', '').replace(' ', '').strip()
                        try:
                            km_value = int(float(km_value))
                        except:
                            km_value = km
                    ws.cell(row=26, column=target_col).value = km_value
                except Exception as e:
                    st.warning(f"      ⚠️ Erreur kilométrage L26: {e}")
        
        # Injection consommation mixte (ligne 23)
        consommation = vehicule_data.get('consommation', '') or vehicule_data.get('conso', '') or vehicule_data.get('consommation_mixte', '')
        if consommation and str(consommation).strip() not in ['', 'nan', 'None']:
            try:
                conso_text = str(consommation).strip()
                
                # Si déjà avec unité, garder tel quel
                if any(unit in conso_text.lower() for unit in ['wh/km', 'kwh/100km', 'l/100km', 'wh', 'kwh', 'l/']):
                    # Déjà formaté
                    pass
                # Si c'est juste un nombre, ajouter l'unité appropriée
                elif conso_text.replace('.', '').replace(',', '').isdigit():
                    conso_num = float(conso_text.replace(',', '.'))
                    vehicule_nom = str(vehicule_data.get('vehicule', '')).lower()
                    
                    # Détection automatique d'unité selon la valeur et le véhicule
                    if any(keyword in vehicule_nom for keyword in ['electric', 'électrique', 'ev', 'id.', 'e-']):
                        # Véhicule électrique
                        if conso_num > 50:  # Probablement Wh/km (ex: 152.80)
                            conso_text += " Wh/km"
                        else:  # Probablement kWh/100km (ex: 16.8)
                            conso_text += " kWh/100km"
                    else:
                        # Véhicule thermique
                        conso_text += " L/100km"
                        
                ws.cell(row=23, column=target_col).value = conso_text
            except Exception as e:
                st.warning(f"      ⚠️ Erreur consommation L23: {e}")
        
        # Injection puissance véhicule (ligne 15)
        puissance = vehicule_data.get('puissance', '') or vehicule_data.get('power', '') or vehicule_data.get('ch', '')
        if puissance and str(puissance).strip() not in ['', 'nan', 'None']:
            try:
                ws.cell(row=15, column=target_col).value = str(puissance).strip()
            except Exception as e:
                st.warning(f"      ⚠️ Erreur puissance L15: {e}")
        
        
        # Injection AND - Amortissements non déductibles / mois (ligne 49)
        and_value = vehicule_data.get('and', '') or vehicule_data.get('amortissement', '') or vehicule_data.get('amort', '')
        if and_value and str(and_value).strip() not in ['', 'nan', 'None']:
            try:
                and_numeric = float(and_value)
                ws.cell(row=49, column=target_col).value = and_numeric
            except (ValueError, TypeError):
                # Si ce n'est pas un nombre, injecter tel quel
                ws.cell(row=49, column=target_col).value = str(and_value)
        
        # Injection TVS (ligne 43)
        tvs_value = vehicule_data.get('tvs', '') or vehicule_data.get('TVS', '') or vehicule_data.get('taxe_vehicule', '')
        if tvs_value and str(tvs_value).strip() not in ['', 'nan', 'None']:
            try:
                tvs_numeric = float(tvs_value)
                ws.cell(row=43, column=target_col).value = tvs_numeric
            except (ValueError, TypeError):
                # Si ce n'est pas un nombre, injecter tel quel
                ws.cell(row=43, column=target_col).value = str(tvs_value)
        
        # Traitement des autres champs
        for data_key, value in vehicule_data.items():
            if data_key == 'vehicule':  # Déjà traité ci-dessus
                continue
                
            if value is not None and str(value).strip() not in ['', 'nan', 'None']:
                
                # Chercher le champ correspondant dans la matrice
                possible_fields = mapping_matrice.get(data_key, [data_key])
                
                for field_name in possible_fields:
                    # Recherche flexible du champ
                    matching_row = None
                    for excel_field, row_num in champs_excel.items():
                        if field_name.lower() in excel_field.lower():
                            matching_row = row_num
                            break
                    
                    if matching_row and matching_row != 28:  # Exclure la ligne 28
                        try:
                            if isinstance(value, (int, float)):
                                ws.cell(row=matching_row, column=target_col).value = float(value)
                            else:
                                ws.cell(row=matching_row, column=target_col).value = str(value)
                            
                            break
                        except Exception as e:
                            st.warning(f"      ⚠️ Erreur {data_key}: {e}")
                    elif matching_row == 28:
                        pass  # Ligne 28 exclue
    
    st.success(f"✅ {max_vehicules} véhicule(s) injectés dans la matrice TCO")
    
    output = BytesIO()
    template_wb.save(output)
    output.seek(0)
    return output

def export_matrice_tco_simple(df: pd.DataFrame) -> BytesIO:
    """Crée une matrice TCO simple (format transposé)"""
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Matrice TCO"
    
    # Champs de la matrice
    champs = [
        "Véhicule Complet", "Marque", "Modèle", "Loueur", "Prix Catalogue TTC", "Prix Batterie", 
        "1er Loyer", "Loyer Mensuel TTC", "Durée (mois)", "Kilométrage", "CO2 (g/km)", 
        "TVS Annuelle", "AND Mensuel", "Prestations", "Date Offre"
    ]
    
    # Créer les labels (colonne A)
    for row_idx, champ in enumerate(champs, 1):
        cell = ws.cell(row=row_idx, column=1, value=champ)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
    
    # Remplir les véhicules (max 10 colonnes)
    max_vehicules = min(len(df), 10)
    
    for col_idx in range(max_vehicules):
        col_num = col_idx + 2
        vehicule_data = df.iloc[col_idx]
        
        # Véhicule complet
        vehicule_complet = str(vehicule_data.get('vehicule', f'Véhicule {col_idx+1}'))
        ws.cell(row=1, column=col_num, value=vehicule_complet[:25])
        
        # Séparer marque et modèle
        parts = vehicule_complet.strip().split(' ', 1)
        marque = parts[0] if parts else ''
        modele = parts[1] if len(parts) > 1 else ''
        
        # Données
        ws.cell(row=2, column=col_num, value=marque)  # Marque
        ws.cell(row=3, column=col_num, value=modele)  # Modèle
        ws.cell(row=4, column=col_num, value=vehicule_data.get('loueur', ''))
        ws.cell(row=5, column=col_num, value=vehicule_data.get('prix_catalogue', ''))
        ws.cell(row=6, column=col_num, value=vehicule_data.get('prix_batterie', ''))  # Prix Batterie
        ws.cell(row=7, column=col_num, value=vehicule_data.get('premier_loyer', ''))  # 1er Loyer
        ws.cell(row=8, column=col_num, value=vehicule_data.get('loyers_v1', ''))
        
        # Loi de roulage - séparer durée et km
        loi = str(vehicule_data.get('loi_de_roulage', ''))
        if '/' in loi:
            duree, km = loi.split('/', 1)
            ws.cell(row=9, column=col_num, value=duree)  # Durée (mois)
            ws.cell(row=10, column=col_num, value=f"{km}000" if km.isdigit() else km)  # Kilométrage
        
        ws.cell(row=11, column=col_num, value=vehicule_data.get('co2', ''))  # CO2 (g/km)
        ws.cell(row=12, column=col_num, value=vehicule_data.get('tvs', ''))  # TVS Annuelle
        
        ws.cell(row=13, column=col_num, value=vehicule_data.get('and', ''))  # AND Mensuel
        ws.cell(row=14, column=col_num, value=vehicule_data.get('presta', ''))  # Prestations
        ws.cell(row=15, column=col_num, value=vehicule_data.get('date_offre', ''))  # Date Offre
    
    # Ajuster les largeurs
    ws.column_dimensions['A'].width = 20
    for col_idx in range(2, max_vehicules + 2):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 15
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def main():
    # Interface simplifiée - juste la page d'extraction
    extraction_page()

def extraction_page():
    """Page d'extraction de devis"""
    st.subheader("📥 Extraction de Devis")
    
    # Vérification API pour cette page
    api_key = os.getenv("GOOGLE_API_KEY")
    if not api_key:
        st.error("🚫 Clé API Google Gemini non trouvée dans .env")
        st.code("GOOGLE_API_KEY=votre_clé_api", language="bash")
        st.info("💡 Obtenez votre clé API gratuite sur https://aistudio.google.com/app/apikey")
        st.stop()
    
    # ÉTAPE 1: Choisir le template de destination (AVANT extraction)
    st.write("### Étape 1 : Choisir la destination")
    st.write("Choisissez votre template de destination AVANT l'extraction pour une injection automatique")
    
    col_template1, col_template2 = st.columns(2)
    
    with col_template1:
        excel_suivi_template = st.file_uploader(
            "Excel de Suivi (template)",
            type="xlsx",
            help="Votre fichier Excel de suivi existant",
            key="pre_suivi_template"
        )
        if excel_suivi_template:
            st.success("✅ Template Excel Suivi chargé")
    
    with col_template2:
        matrice_tco_template = st.file_uploader(
            "Matrice TCO (template)", 
            type="xlsx",
            help="Votre matrice TCO existante",
            key="pre_matrice_template"
        )
        if matrice_tco_template:
            st.success("✅ Template Matrice TCO chargé")
    
    st.write("---")
    
    # ÉTAPE 2: Upload des PDFs
    st.write("### Étape 2 : Charger les PDFs à extraire")
    pdfs = st.file_uploader(
        "Chargez vos devis PDF", 
        type="pdf", 
        accept_multiple_files=True,
        help="Supports PDFs texte ET scannés avec OCR"
    )
    
    if pdfs:
        st.info(f"🔄 Traitement de {len(pdfs)} PDF(s)")
        
        # Créer l'extracteur
        extractor = DevisExtractorSimplifie(api_key)
        
        if extractor.ocr_disponible:
            st.info(f"🔍 OCR activé - Tesseract {extractor.tesseract_version}")
        else:
            st.warning("⚠️ OCR non disponible - Extraction texte uniquement")
        
        # Traitement
        button_text = "🚀 Extraire les devis"
        if excel_suivi_template or matrice_tco_template:
            button_text += " et injecter automatiquement"
        
        if st.button(button_text, type="primary"):
            start_time = time.time()
            records = process_pdfs_simplifie(extractor, pdfs)
            processing_time = time.time() - start_time
            
            st.success(f"✅ Traitement terminé en {processing_time:.1f}s")
            
            if not records:
                st.warning("⚠️ Aucun devis traité")
                st.stop()
            
            # DataFrame
            df = pd.DataFrame(records)
            
            # Affichage des résultats
            st.subheader("🔍 Données extraites")
            
            col1, col2, col3 = st.columns(3)
            col1.metric("📄 PDFs traités", len(pdfs))
            col2.metric("✅ Devis extraits", len(df))
            
            # Métriques financières
            try:
                if 'loyers_v1' in df.columns:
                    loyers_numeric = pd.to_numeric(df['loyers_v1'], errors='coerce')
                    avg_loyer = loyers_numeric.dropna().mean()
                    if not pd.isna(avg_loyer):
                        col3.metric("💰 Loyer moyen", f"{avg_loyer:.0f}€")
            except:
                pass
            
            # Tableau des données
            st.write("📊 **Résultats d'extraction :**")
            for i, row in df.iterrows():
                with st.expander(f"📄 Devis {i+1} - {row.get('vehicule', 'N/A')} ({row.get('loueur', 'N/A')})"):
                    col1, col2 = st.columns(2)
                    with col1:
                        # Prix catalogue - toujours afficher
                        prix_cat = row.get('prix_catalogue', 'N/A')
                        if prix_cat is not None and str(prix_cat).strip() not in ['', 'N/A', 'None']:
                            st.write(f"💳 **Prix catalogue:** {prix_cat}€")
                        else:
                            st.write(f"💳 **Prix catalogue:** Non extrait")
                        
                        # Loyer mensuel - toujours afficher  
                        loyer = row.get('loyers_v1', 'N/A')
                        if loyer is not None and str(loyer).strip() not in ['', 'N/A', 'None']:
                            st.write(f"💰 **Loyer mensuel:** {loyer}€")
                        else:
                            st.write(f"💰 **Loyer mensuel:** Non extrait")
                        
                        # Loi de roulage
                        if row.get('loi_de_roulage'):
                            st.write(f"📋 **Loi de roulage:** {row['loi_de_roulage']}")
                    with col2:
                        if row.get('date_offre'):
                            st.write(f"📅 **Date offre:** {row['date_offre']}")
                        if row.get('co2'):
                            st.write(f"🌿 **CO2:** {row['co2']} g/km")
                        if row.get('presta'):
                            st.write(f"🛡️ **Prestations:** {row['presta']}")
                    
                    # Debug : afficher toutes les données extraites (section repliable)
                    st.write("---")
                    st.write("🔧 **Debug - Toutes les données extraites :**")
                    debug_cols = st.columns(2)
                    items = list(row.items())
                    mid = len(items) // 2
                    
                    with debug_cols[0]:
                        for key, value in items[:mid]:
                            st.write(f"**{key}:** {value}")
                    
                    with debug_cols[1]:
                        for key, value in items[mid:]:
                            st.write(f"**{key}:** {value}")
            
            # INJECTION AUTOMATIQUE ou création selon templates pré-chargés
            st.subheader("📤 Injection et Export des données")
            
            # Variables pour stocker les outputs
            injection_results = []
            
            # 1. Injection automatique dans Excel Suivi si template pré-chargé
            if excel_suivi_template:
                st.write("### 💉 Injection automatique dans Excel de Suivi")
                try:
                    template_wb = openpyxl.load_workbook(excel_suivi_template)
                    output_inject = inject_in_excel_suivi(df, template_wb)
                    filename_inject = f"Suivi_Injecte_{datetime.today():%Y%m%d_%H%M}.xlsx"
                    
                    st.success("✅ Injection dans Excel Suivi réussie !")
                    
                    st.download_button(
                        "📥 Télécharger Excel Suivi Mis à Jour",
                        data=output_inject.read(),
                        file_name=filename_inject,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        type="primary"
                    )
                    injection_results.append("Excel Suivi")
                    
                except Exception as e:
                    st.error(f"❌ Erreur injection Excel Suivi: {e}")
            
            # 2. Injection automatique dans Matrice TCO si template pré-chargé
            if matrice_tco_template:
                st.write("### 💉 Injection automatique dans Matrice TCO")
                try:
                    template_wb = openpyxl.load_workbook(matrice_tco_template)
                    output_inject = inject_in_matrice_tco(df, template_wb)
                    filename_inject = f"Matrice_Injectee_{datetime.today():%Y%m%d_%H%M}.xlsx"
                    
                    st.success("✅ Injection dans Matrice TCO réussie !")
                    
                    st.download_button(
                        "📥 Télécharger Matrice TCO Mise à Jour",
                        data=output_inject.read(),
                        file_name=filename_inject,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        type="primary"
                    )
                    injection_results.append("Matrice TCO")
                    
                except Exception as e:
                    st.error(f"❌ Erreur injection Matrice TCO: {e}")
            
            # 3. Si aucun template pré-chargé, proposer création de nouveaux fichiers
            if not excel_suivi_template and not matrice_tco_template:
                st.write("### 🆕 Création de nouveaux fichiers")
                st.info("💡 Aucun template pré-chargé - Création de nouveaux fichiers")
                
                col1, col2 = st.columns(2)
                
                with col1:
                    st.write("**📊 Excel de Suivi**")
                    if st.button("📥 Créer Excel Suivi", use_container_width=True):
                        output_suivi = export_excel_suivi(df)
                        filename_suivi = f"Suivi_Devis_{datetime.today():%Y%m%d_%H%M}.xlsx"
                        
                        st.download_button(
                            "📥 Télécharger Excel Suivi",
                            data=output_suivi.read(),
                            file_name=filename_suivi,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                
                with col2:
                    st.write("**🧮 Matrice TCO**")
                    if st.button("📥 Créer Matrice TCO", use_container_width=True):
                        output_matrice = export_matrice_tco_simple(df)
                        filename_matrice = f"Matrice_TCO_{datetime.today():%Y%m%d_%H%M}.xlsx"
                        
                        st.download_button(
                            "📥 Télécharger Matrice TCO",
                            data=output_matrice.read(),
                            file_name=filename_matrice,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
            
            # 4. Options supplémentaires supprimées pour interface plus propre
            elif injection_results:
                pass  # Injection automatique terminée silencieusement

if __name__ == "__main__":
    extraction_page() 
